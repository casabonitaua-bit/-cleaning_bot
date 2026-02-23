import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database import get_db

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header_row(ws, headers: list[str], row: int):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws):
    for col in ws.columns:
        try:
            width = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = width
        except AttributeError:
            pass


async def excel_city_base(city: str) -> io.BytesIO:
    async with get_db() as db:
        rows = await db.fetch(
            """
            SELECT up.full_name, up.phone, u.username,
                   up.age, up.rating, up.total_shifts, up.confirmed_shifts,
                   up.refused_shifts, up.ignored_shifts,
                   up.consecutive_failures, up.is_active
            FROM user_profiles up
            JOIN users u ON u.telegram_id = up.telegram_id
            WHERE up.city = $1
            ORDER BY up.full_name
            """,
            city
        )

    wb = Workbook()
    ws = wb.active
    ws.title = city

    ws.append([f"База сотрудников — {city}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "ФИО", "Телефон", "Telegram", "Возраст", "Рейтинг",
        "Всего смен", "Подтверждено", "Отказов", "Игноров",
        "Подряд провалов", "Активен"
    ]
    _write_header_row(ws, headers, 3)

    for r in rows:
        ws.append([
            r["full_name"],
            r["phone"],
            f"@{r['username']}" if r["username"] else "—",
            r["age"],
            r["rating"],
            r["total_shifts"],
            r["confirmed_shifts"],
            r["refused_shifts"],
            r["ignored_shifts"],
            r["consecutive_failures"],
            "✅" if r["is_active"] else "🚫",
        ])

    _autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def excel_shift_report(shift_id: int) -> io.BytesIO:
    async with get_db() as db:
        shift = await db.fetchrow(
            "SELECT * FROM shifts WHERE id = $1",
            shift_id
        )

        rows = await db.fetch(
            """
            SELECT up.full_name, up.phone, up.rating,
                   sm.member_type, sm.status, sm.position,
                   sr.worked, sr.decline_reason
            FROM shift_members sm
            JOIN user_profiles up ON sm.telegram_id = up.telegram_id
            LEFT JOIN shift_results sr
                ON sr.shift_id = sm.shift_id AND sr.telegram_id = sm.telegram_id
            WHERE sm.shift_id = $1
            ORDER BY sm.member_type, sm.position
            """,
            shift_id
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Смена {shift_id}"

    # Если смена не найдена — не падаем, а формируем понятный файл
    if not shift:
        ws.append([f"Смена #{shift_id} не найдена"])
        ws["A1"].font = Font(bold=True, size=12)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # shift: id=0, city=1, date=2, address=3 (оставляю совместимость с твоей логикой)
    title = f"Смена #{shift_id} | {shift[1]} | {shift[2]} | {shift[3]}"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "ФИО", "Телефон", "Рейтинг",
        "Тип", "Статус", "Позиция", "Отработал", "Причина"
    ]
    _write_header_row(ws, headers, 3)

    for r in rows:
        worked = "—"
        if r["worked"] == 1:
            worked = "Да"
        elif r["worked"] == 0:
            worked = "Нет"

        ws.append([
            r["full_name"],
            r["phone"],
            r["rating"],
            "Основа" if r["member_type"] == "main" else "Резерв",
            r["status"],
            r["position"],
            worked,
            r["decline_reason"] or "",
        ])

    _autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf